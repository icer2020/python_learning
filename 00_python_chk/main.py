#!/usr/bin/python3.6 -u
# -*- coding: utf-8 -*-

#  Copyright (c) 2022-2022. WAYON All rights Reserved.

import argparse
import datetime
import openpyxl as ox
import os
import re
import shutil
import sys
import time
import utils_wo as ut

"""
    File        test.py
    Author      Guoqiang
    Date        2022/7/11 11:00
    Describe    basic test
    Version     v0.1
    History     v0.1: 2022/7/11 11:00
"""

# Version history
ver = 'V0.1'
ver_date = 'July 11, 2022'
ver_des = 'initial version.'
ver_detail_des = '''initial version '''
scr_des = "Basic test"
scr_des_detail = '''
Description: Basic test 
'''
starttime = time.time()

ut.ver = ver
ut.ver_date = ver_date
ut.ver_des = ver_des
ut.ver_detail_des = ver_detail_des
ut.scr_des = scr_des
ut.starttime = starttime


# ------------------------------------------------------------
# class section
# ------------------------------------------------------------
class ExcelParse:
    def __init__(self, file):
        self.file = file
        self.wb = ox.load_workbook(self.file, data_only=True)
        # sheets = self.wb[]
        # self.sheet = self.wb.worksheets[0]
        self.ws = self.wb.worksheets[0]

    # 获取表格的总行数和总列数
    def get_rows_clos_num(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    def copy_ws(self):
        # self.wb.copy_worksheet
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def get_cell_value(self, row, column):
        cellvalue = self.ws.cell(row=row, column=column).value
        return cellvalue

    # 获取某列的所有值
    def get_col_values(self, column):
        rows = self.ws.max_row
        columndata = []
        for i in range(1, rows + 1):
            cellvalue = self.ws.cell(row=i, column=column).value
            columndata.append(cellvalue)
        return columndata

    # 获取某行所有值
    def get_row_values(self, row):
        columns = self.ws.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = self.ws.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    # 设置某个单元格的值
    def set_cell_value(self, row, column, cellvalue):
        try:
            self.ws.cell(row=row, column=column).value = cellvalue
            self.wb.save(self.file)
        except IOError:
            self.ws.cell(row=row, column=column).value = "writefail"
            self.wb.save(self.file)

# ------------------------------------------------------------
# function section
# ------------------------------------------------------------
def parseargu():
    global ver
    global ver_date
    global ver_des
    post_msg = ver.ljust(18) + ver_des.center(40) + ver_date.rjust(18)
    parser = argparse.ArgumentParser(
        formatter_class=argparse.RawTextHelpFormatter,
        description=scr_des_detail, epilog=post_msg)

    parser.add_argument('-i', '--in_file', required=False,
                        help='Input file',
                        default='./user_input.rpt')
    parser.add_argument('-o', '--out_file', required=False,
                        help='Output file',
                        default='./user_output.rpt')
    parser.add_argument('-sta', '--status', type=str, choices=['RUN', 'DEBUG'],
                        help=argparse.SUPPRESS)
    args_l = parser.parse_args()
    # print('args_l', args_l)
    return args_l

def main_run(i_file, o_file):
    fo = open(o_file, 'w')
    cnt = 1
    for i in open(i_file).readlines():
        outline=('{:<4s}{:>03d} {:<s}'.format("##", cnt, i.rstrip()))
        fo.writelines(outline+'\n')
        cnt+=1
    fo.close()

# ------------------------------------------------------------
# Running  main
# ------------------------------------------------------------
if __name__ == '__main__':
    # run status:  "DEBUG" || "INFO" || "WARNING" || "ERROR" || "CRITICAL" || "RUN":
    ut.RUN_STATUS = "RUN"

    # additional meg in header for current scripts
    ut.header(add_msg='')

    rpt_dir = 'report'
    in_dir = 'input'
    out_date = str(datetime.datetime.now().strftime('%m%d'))
    out_dir = "mem_gen_" + out_date

    # path delimiter
    # dt = '/'

    args = parseargu()

    main_run(i_file=args.in_file, o_file=args.out_file)
    ut.footer()